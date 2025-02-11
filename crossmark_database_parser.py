# INTEL CONFIDENTIAL
# Copyright 2015 Intel Corporation All Rights Reserved.
#
# The source code contained or described herein and all documents related
# to the source code ("Material") are owned by Intel Corporation or its
# suppliers or licensors. Title to the Material remains with Intel Corp-
# oration or its suppliers and licensors. The Material may contain trade
# secrets and proprietary and confidential information of Intel Corpor-
# ation and its suppliers and licensors, and is protected by worldwide
# copyright and trade secret laws and treaty provisions. No part of the
# Material may be used, copied, reproduced, modified, published, uploaded,
# posted, transmitted, distributed, or disclosed in any way without
# Intel's prior express written permission.
#
# No license under any patent, copyright, trade secret or other intellect-
# ual property right is granted to or conferred upon you by disclosure or
# delivery of the Materials, either expressly, by implication, inducement,
# estoppel or otherwise. Any license under such intellectual property
# rights must be express and approved by Intel in writing.

# Author: Ronny Z. Valtonen
# Date Created: 07/17/2023
# Purpose: To assist on parsing files created by Crossmark Debug Tool.

#####################################
# BASIC PREP                        #
# pip install openpyxl (3.0.10)     #
# pip install xlsxwriter (3.0.3)    #
# pip install pandas (1.4.2)        #
# python 3.9.12                     #
# Linux: sudo pacman -S tk          #
# MacOS: brew install python-tk     #
#####################################

# Program
import os
import xlsxwriter
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook


# UI
from tkinter import filedialog
from tkinter import *


# Parses the initial Crossmark test scores.
def parse_performance(file, sheet):
    print("Found initial Crossmark performance csv, beginning data collection.")

    # Get run folder
    split_data = file.split("\\")
    my_file = split_data[-2]

    # Collect the data here
    # Create a worksheet to write to.
    my_worksheet = sheet.add_worksheet(my_file)

    # with open(file, newline = '') as csvfile:
    #     main_reader = csv.reader(csvfile, delimiter = ' ', quotechar = '|')
    #     for row in main_reader:
    #         print(', '.join(row))

    with open(file) as main_scores:
        reader = csv.reader(main_scores)
        rows = list(reader)
        # Printing main scores
        print("Printing Main Scores...")
        # print(rows)
    
    # Get scores
    real_row_1 = rows[1]
    real_row_2 = rows[6]
    real_row_3 = rows[14]
    real_row_4 = rows[20]
    crossmark_score = real_row_1[4]
    productivity_score = real_row_2[4]
    creativity_score = real_row_3[4]
    responsiveness_score = real_row_4[4]

    # Once done, parse the subtest csv file.
    print("Grabbing subtests")

    # Convert the file into a string, remove the last scores.csv characters, and add in the desired csv file we want next.
    convert_to_string = str(file)
    previous_directory = (convert_to_string[0:-10])
    previous_directory += "measure_performance.csv"
    print(previous_directory)

    with open(previous_directory) as sub_scores:
        reader_second = csv.reader(sub_scores)
        sub_rows = list(reader_second)
        # Printed sub scores
        print("Printing Sub-Test Scores")
        # print(sub_rows)

    # Get scores
    sub_row_1 = sub_rows[23]
    sub_row_2 = sub_rows[24]
    sub_row_3 = sub_rows[25]
    sub_row_4 = sub_rows[26]
    sub_row_5 = sub_rows[27]
    sub_row_6 = sub_rows[28]
    sub_row_7 = sub_rows[29]
    sub_row_8 = sub_rows[30]
    sub_row_9 = sub_rows[31]
    sub_row_10 = sub_rows[32]
    sub_row_11 = sub_rows[33]
    sub_row_12 = sub_rows[34]
    sub_row_13 = sub_rows[35]
    sub_row_14 = sub_rows[36]
    sub_row_15 = sub_rows[37]
    sub_row_16 = sub_rows[38]
    sub_row_17 = sub_rows[39]
    sub_row_18 = sub_rows[40]
    sub_row_19 = sub_rows[41]
    sub_row_20 = sub_rows[42]
    sub_row_21 = sub_rows[43]
    sub_row_22 = sub_rows[44]
    print(sub_rows[23])

    # Grab the desired values
    zstd_uncompress_legacy = sub_row_1[-2]
    random_read = sub_row_2[-2]
    black_scholes_serial= sub_row_3[-2]
    string_search = sub_row_4[-2]
    random_write = sub_row_5[-2]
    object_detection = sub_row_6[-2]
    ef_face_recognition = sub_row_7[-2]
    zstd_compress_legacy = sub_row_8[-2]
    zstd_uncompress_streaming = sub_row_9[-2]
    fdt_by_medianflow_tracker = sub_row_10[-2]
    black_scholes_parallel = sub_row_11[-2]
    create_sqlite_blob = sub_row_12[-2]
    video_colorization = sub_row_13[-2]
    external_sort = sub_row_14[-2]
    chacha20_encrypt_openssl = sub_row_15[-2]
    colorization = sub_row_16[-2]
    hdr_stitch = sub_row_17[-2]
    aes_gcm_encrypt_mt = sub_row_18[-2]
    memory_workload = sub_row_19[-2]
    chacha20_decrypt_openssl = sub_row_20[-2]
    gzip_compress = sub_row_21[-2]
    gzip_uncompress = sub_row_22[-2]

    my_worksheet.write('A2', crossmark_score)
    my_worksheet.write('B2', productivity_score)
    my_worksheet.write('C2', creativity_score)
    my_worksheet.write('D2', responsiveness_score)
    my_worksheet.write('E2', zstd_uncompress_legacy)
    my_worksheet.write('F2', random_read)
    my_worksheet.write('G2', black_scholes_serial)
    my_worksheet.write('H2', string_search)
    my_worksheet.write('I2', random_write)
    my_worksheet.write('J2', object_detection)
    my_worksheet.write('K2', ef_face_recognition)
    my_worksheet.write('L2', zstd_compress_legacy)
    my_worksheet.write('M2', zstd_uncompress_streaming)
    my_worksheet.write('N2', fdt_by_medianflow_tracker)
    my_worksheet.write('O2', black_scholes_parallel)
    my_worksheet.write('P2', create_sqlite_blob)
    my_worksheet.write('Q2', video_colorization)
    my_worksheet.write('R2', external_sort)
    my_worksheet.write('S2', chacha20_encrypt_openssl)
    my_worksheet.write('T2', colorization)
    my_worksheet.write('U2', hdr_stitch)
    my_worksheet.write('V2', aes_gcm_encrypt_mt)
    my_worksheet.write('W2', memory_workload)
    my_worksheet.write('X2', chacha20_decrypt_openssl)
    my_worksheet.write('Y2', gzip_compress)
    my_worksheet.write('Z2', gzip_uncompress)

    
# Selects the appropriate files within directory.
def pick_file(window, selected_file, workbook):
    if "scores.csv" in selected_file:
        parse_performance(selected_file, workbook)


# Merge all of the sheets together into one excel document.
def combine_results():
    try:
        df = pd.concat(pd.read_excel('debug.xlsx', sheet_name = None), ignore_index = False)
        pd.options.display.precision = 3
        df.to_excel("combined_data.xlsx")
        transposer()
        #os.system("start EXCEL.exe combined_data.xlsx")
        

    except:
        pass

# Converts vertical layout into horizontal.
def transposer():
    print("Transposing...")
    
    # Open the workbook
    wb = load_workbook("combined_data.xlsx")

    benchmark = {
        "A1": "Iteration",
        "C1": "Crossmark Overall Score",
        "D1": "Productivity Score",
        "E1": "Creativity Score",
        "F1": "Responsiveness Score",
        "G1": "zstd_uncompress_legacy",
        "H1": "random_read",
        "I1": "black_scholes_serial",
        "J1": "string_search",
        "K1": "random_write",
        "L1": "object_detection",
        "M1": "ef_face_recognition",
        "N1": "zstd_compress_legacy",
        "O1": "zstd_uncompress_streaming",
        "P1": "fdt_by_medianflow_tracker",
        "Q1": "black_scholes_parallel",
        "R1": "create_sqlite_blob",
        "S1": "video_colorization",
        "T1": "external_sort",
        "U1": "chacha20_encrypt_openssl",
        "V1": "colorization",
        "W1": "hdr_stitch",
        "X1": "aes_gcm_encrypt_mt",
        "Y1": "memory_workload",
        "Z1": "chacha20_decrypt_openssl",
        "AA1": "gzip_compress",
        "AB1": "gzip_uncompress",
    }

    print("Printing keys")
    x = benchmark.keys()
    print(x)

    # Change cell values to approriate terms.
    sheet = wb.worksheets[0]
    columns = sheet["A1"].value = "Iteration"
    columns = sheet["C1"].value = "Crossmark Overall Score"
    columns = sheet["D1"].value = "Productivity Score"
    columns = sheet["E1"].value = "Creativity Score"
    columns = sheet["F1"].value = "Responsiveness Score"
    columns = sheet["G1"].value = "zstd_uncompress_legacy"
    columns = sheet["H1"].value = "random_read"
    columns = sheet["I1"].value = "black_scholes_serial"
    columns = sheet["J1"].value = "string_search"
    columns = sheet["K1"].value = "random_write"
    columns = sheet["L1"].value = "object_detection"
    columns = sheet["M1"].value = "ef_face_recognition"
    columns = sheet["N1"].value = "zstd_compress_legacy"
    columns = sheet["O1"].value = "zstd_uncompress_streaming"
    columns = sheet["P1"].value = "fdt_by_medianflow_tracker"
    columns = sheet["Q1"].value = "black_scholes_parallel"
    columns = sheet["R1"].value = "create_sqlite_blob"
    columns = sheet["S1"].value = "video_colorization"
    columns = sheet["T1"].value = "external_sort"
    columns = sheet["U1"].value = "chacha20_encrypt_openssl"
    columns = sheet["V1"].value = "colorization"
    columns = sheet["W1"].value = "hdr_stitch"
    columns = sheet["X1"].value = "aes_gcm_encrypt_mt"
    columns = sheet["Y1"].value = "memory_workload"
    columns = sheet["Z1"].value = "chacha20_decrypt_openssl"
    columns = sheet["AA1"].value = "gzip_compress"
    columns = sheet["AB1"].value = "gzip_uncompress"
    wb.save("combined_data.xlsx")
    print("Transposing complete.")

# Program driver.
def main():
    # Declare a window.
    window = Tk()
    print("Prompting user with File Explorer")

    window.filename = filedialog.askdirectory(initialdir= "/", title = "Select Debug Folder")
    window.geometry("750x270")

    # Initialize label
    Label(window, text = "Parsing Data, Please Wait... \n Created by Ronny V.", font=('Helvetica 20 bold')).pack(pady=20)
    window.after(1000, lambda: window.destroy())
    window.mainloop()

    # Check if the files exist, if they do, delete them so that we don't have to overwrite them.
    if os.path.exists('debug.xlsx'):
        os.remove('debug.xlsx')

    if os.path.exists('combined_data.xlsx'):
        os.remove('combined_data.xlsx')

    workbook = xlsxwriter.Workbook('debug.xlsx', {'constant_memory': True})

    for root, dirs, files in os.walk(window.filename):
        for file in files:
            try:
                selected_file = os.path.join(root, file)
                pick_file(window, selected_file, workbook)

            except:
                pass

    workbook.close()



# Runs program in correct order.
if __name__ == "__main__":
    print("Starting program")
    main()
    print("Combining results.")
    combine_results()
    print("Parsing complete.")
